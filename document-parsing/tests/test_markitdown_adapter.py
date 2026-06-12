"""Unit tests for the MarkItDown spreadsheet adapter and routing parser."""

import io

import pytest
from openpyxl import Workbook

from document_parsing.config import ParserConfig
from document_parsing.factory import create_document_parser
from document_parsing.markitdown_adapter import (
    MarkItDownConversionError,
    MarkItDownDocumentParser,
    is_markitdown_format,
)
from document_parsing.routing_adapter import RoutingDocumentParser


def _build_xlsx() -> bytes:
    """Return bytes of a tiny single-sheet workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rev Trk'
    ws.append(['Month', 'AtNeed', 'PreNeed'])
    ws.append(['Jan', 1200, 800])
    ws.append(['Feb', 1500, 950])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _StubParser:
    """Primary-parser stub that records whether it was called."""

    def __init__(self) -> None:
        self.called_with: str | None = None

    def parse(self, file_content: bytes, filename: str) -> list[dict]:
        self.called_with = filename
        return [{'type': 'NarrativeText', 'text': 'docling', 'metadata': {}}]

    def health_check(self) -> dict:
        return {'status': 'healthy', 'parser': 'stub'}


@pytest.mark.unit
@pytest.mark.parametrize(
    ('filename', 'expected'),
    [
        ('CSNH Rev Trk.xlsx', True),
        ('legacy.xls', True),
        ('data.csv', True),
        ('REPORT.XLSX', True),
        ('doc.pdf', False),
        ('slides.pptx', False),
        ('image.png', False),
    ],
)
def test_is_markitdown_format(filename: str, expected: bool) -> None:
    assert is_markitdown_format(filename) is expected


@pytest.mark.unit
def test_markitdown_parses_xlsx_to_table_element() -> None:
    parser = MarkItDownDocumentParser()
    elements = parser.parse(_build_xlsx(), 'CSNH Rev Trk.xlsx')

    assert len(elements) == 1
    el = elements[0]
    assert el['type'] == 'Table'
    # .xlsx now goes through the bounded-memory openpyxl streaming path
    assert el['metadata']['parser'] == 'openpyxl-stream'
    assert el['metadata']['filename'] == 'CSNH Rev Trk.xlsx'
    # text_as_html mirrors the markdown content for downstream consumers
    assert el['metadata']['text_as_html'] == el['text']
    # data made it through
    assert 'Month' in el['text'] and 'AtNeed' in el['text'] and '1200' in el['text']


@pytest.mark.unit
def test_xlsx_streaming_renders_multiple_sheets_and_escapes_pipes() -> None:
    """Streaming path renders every sheet and escapes ``|`` in cell values."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'AtNeed'
    ws1.append(['Region', 'Note'])
    ws1.append(['North', 'a|b'])  # pipe must be escaped, not break the table
    ws2 = wb.create_sheet('PreNeed')
    ws2.append(['Region', 'Total'])
    ws2.append(['South', 42])
    buf = io.BytesIO()
    wb.save(buf)

    elements = MarkItDownDocumentParser().parse(buf.getvalue(), 'multi.xlsx')

    assert len(elements) == 1
    text = elements[0]['text']
    assert elements[0]['metadata']['parser'] == 'openpyxl-stream'
    # both sheets present
    assert '## AtNeed' in text and '## PreNeed' in text
    # both rows present
    assert 'North' in text and 'South' in text and '42' in text
    # pipe escaped so it does not split the cell
    assert 'a\\|b' in text


@pytest.mark.unit
def test_xls_legacy_falls_back_to_markitdown() -> None:
    """Legacy ``.xls`` (not OOXML) is delegated to MarkItDown, not openpyxl."""

    class _RecordingConverter:
        def __init__(self) -> None:
            self.called = False

        def convert_stream(self, *_args: object, **_kwargs: object) -> object:
            self.called = True

            class _R:
                text_content = '| col |\n| --- |\n| v |'

            return _R()

    parser = MarkItDownDocumentParser()
    conv = _RecordingConverter()
    parser._converter = conv  # inject so no real .xls parsing is needed

    elements = parser.parse(b'\xd0\xcf\x11\xe0legacy', 'legacy.xls')

    assert conv.called is True  # routed to MarkItDown, not openpyxl
    assert elements[0]['metadata']['parser'] == 'markitdown'


@pytest.mark.unit
def test_corrupt_xlsx_falls_back_to_markitdown() -> None:
    """Bytes openpyxl cannot open as .xlsx fall back to MarkItDown (no crash)."""
    parser = MarkItDownDocumentParser()
    # Not a valid zip/OOXML — openpyxl.load_workbook raises, adapter falls back.
    elements = parser.parse(b'not a real spreadsheet at all', 'broken.xlsx')
    assert isinstance(elements, list)  # must not raise


@pytest.mark.unit
def test_markitdown_concurrency_semaphore_bounds_parallel_parses(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """At most MARKITDOWN_MAX_CONCURRENCY parses run the conversion at once."""
    import threading
    import time

    import document_parsing.markitdown_adapter as mod

    # Rebuild the module-level semaphore at limit=2 for this test.
    monkeypatch.setattr(mod, '_PARSE_SEMAPHORE', threading.BoundedSemaphore(2))

    active = 0
    peak = 0
    lock = threading.Lock()
    xlsx = _build_xlsx()

    orig_render = mod.MarkItDownDocumentParser._render_sheet_markdown

    def _slow_render(sheet: object) -> str:
        nonlocal active, peak
        with lock:
            active += 1
            peak = max(peak, active)
        time.sleep(0.05)  # widen the window so overlap is observable
        try:
            return orig_render(sheet)
        finally:
            with lock:
                active -= 1

    monkeypatch.setattr(
        mod.MarkItDownDocumentParser, '_render_sheet_markdown', staticmethod(_slow_render)
    )

    parser = mod.MarkItDownDocumentParser()
    threads = [threading.Thread(target=parser.parse, args=(xlsx, f'f{i}.xlsx')) for i in range(6)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    assert peak <= 2  # semaphore held concurrency at the configured limit


@pytest.mark.unit
def test_markitdown_health_check_healthy() -> None:
    assert MarkItDownDocumentParser().health_check()['status'] == 'healthy'


@pytest.mark.unit
def test_markitdown_converter_exception_becomes_terminal_error() -> None:
    """A converter that raises is normalized to MarkItDownConversionError.

    This is the terminal-but-non-KB-failing path: the runner marks the doc
    FAILED and continues, exactly like a terminal Docling failure. A ``.csv``
    is used because it still routes through MarkItDown's ``convert_stream``
    (``.xlsx`` now uses the openpyxl streaming path instead).
    """

    class _BoomConverter:
        def convert_stream(self, *_args: object, **_kwargs: object) -> object:
            raise RuntimeError('markitdown could not parse this csv')

    parser = MarkItDownDocumentParser()
    parser._converter = _BoomConverter()  # inject a converter that raises
    with pytest.raises(MarkItDownConversionError):
        parser.parse(b'\xff\xfe garbage', 'broken.csv')


@pytest.mark.unit
def test_router_sends_spreadsheets_to_markitdown_not_primary() -> None:
    stub = _StubParser()
    router = RoutingDocumentParser(primary_parser=stub)

    elements = router.parse(_build_xlsx(), 'CSNH Rev Trk.xlsx')

    # primary (Docling stub) must NOT have been called for a spreadsheet
    assert stub.called_with is None
    # routed to the tabular (MarkItDown) parser — xlsx uses the openpyxl stream
    assert elements[0]['metadata']['parser'] == 'openpyxl-stream'


@pytest.mark.unit
def test_router_delegates_non_spreadsheets_to_primary() -> None:
    stub = _StubParser()
    router = RoutingDocumentParser(primary_parser=stub)

    elements = router.parse(b'%PDF-1.7 ...', 'report.pdf')

    assert stub.called_with == 'report.pdf'
    assert elements[0]['text'] == 'docling'


@pytest.mark.unit
def test_factory_wraps_when_routing_enabled() -> None:
    cfg = ParserConfig(
        docling_serve_url='http://docling:5001',
        route_spreadsheets_to_markitdown=True,
    )
    assert isinstance(create_document_parser(cfg), RoutingDocumentParser)


@pytest.mark.unit
def test_factory_no_wrap_when_routing_disabled() -> None:
    cfg = ParserConfig(
        docling_serve_url='http://docling:5001',
        route_spreadsheets_to_markitdown=False,
    )
    assert not isinstance(create_document_parser(cfg), RoutingDocumentParser)

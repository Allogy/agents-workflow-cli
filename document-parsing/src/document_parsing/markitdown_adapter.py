"""MarkItDown-backed parser for spreadsheets and other tabular formats.

Spreadsheets (``.xlsx``/``.xls``/``.csv``) carry no layout/image content that
needs Docling's VLM + OCR pipeline — they are pure tabular data. Routing them
through Docling Serve is both wasteful and fragile: on prod the ``CSNH ... Rev
Trk`` workbooks blew past the 30-min per-document parse budget (Docling tried
to render embedded objects / un-renderable WMF images), so every spreadsheet
``DoclingParseTimeout``-ed and fail-emptied while still consuming a full slot
for 30 min (see APPS.md kb-orchestration Known Issues, 2026-06-11).

``MarkItDownDocumentParser`` extracts spreadsheets locally via Microsoft's
``markitdown`` (pandas/openpyxl under the hood) in milliseconds — no VLM, no
network, no per-document timeout risk. It returns elements in the same
normalized shape every other adapter produces (``type``/``text``/``metadata``
with ``text_as_html`` on tables), so the downstream HybridChunker is unchanged.

This adapter is intentionally narrow: it handles ONLY tabular formats. The
``RoutingDocumentParser`` dispatches spreadsheets here and delegates everything
else (PDF/DOCX/PPTX/images) to the primary Docling adapter.
"""

import logging
import os
import pathlib
import threading
from io import BytesIO
from typing import Any

from document_parsing.docling_adapter import DoclingConversionError
from document_parsing.utils import sanitize_text

logger = logging.getLogger(__name__)

# Tabular formats handled by MarkItDown instead of Docling. ``.csv`` is
# included for completeness, but the runner's ``try_direct_parse`` already
# short-circuits ``.csv`` before any external/structured parser is reached;
# keeping it here makes the adapter correct even if called directly.
MARKITDOWN_EXTENSIONS: frozenset[str] = frozenset({'.xlsx', '.xls', '.csv'})

# Excel formats that are streamed cell-by-cell via openpyxl (read-only mode)
# instead of MarkItDown's whole-workbook load. ``.csv`` is excluded — it is
# already lightweight and is handled by MarkItDown directly.
_EXCEL_EXTENSIONS: frozenset[str] = frozenset({'.xlsx', '.xls'})


def _markitdown_concurrency() -> int:
    """Return the max number of concurrent spreadsheet parses.

    Reads ``MARKITDOWN_MAX_CONCURRENCY`` (default 1). Spreadsheet parsing
    expands a zip-compressed workbook many-fold in memory, so even with a
    higher document-level concurrency (``PGVECTOR_DOC_CONCURRENCY``) we cap
    how many large workbooks may inflate simultaneously to avoid OOM-killing
    the ingest container (prod 2026-06-12: 4 large ``.xlsx`` expanded at once
    under ``PGVECTOR_DOC_CONCURRENCY=6`` and blew the 4 GB cap, exit 137).
    """
    try:
        return max(1, int(os.environ.get('MARKITDOWN_MAX_CONCURRENCY', '1')))
    except (TypeError, ValueError):
        return 1


# Module-level semaphore shared across ALL parser instances and worker
# threads in a process. This is the memory guard: regardless of how many
# documents the runner parses concurrently, at most ``MARKITDOWN_MAX_CONCURRENCY``
# spreadsheets may inflate in memory at the same time. Sized once at import.
_PARSE_SEMAPHORE = threading.BoundedSemaphore(_markitdown_concurrency())


class MarkItDownConversionError(DoclingConversionError):
    """Terminal MarkItDown conversion failure for a tabular document.

    Subclasses ``DoclingConversionError`` so the ingest runner's existing
    terminal-error handling marks the document FAILED and continues to the
    next one (one bad spreadsheet never fails the whole KB), identical to how
    a terminal Docling failure is treated.
    """


def is_markitdown_format(filename: str) -> bool:
    """Return True when *filename* should be parsed by MarkItDown.

    Args:
        filename: Original filename or S3 key; only the suffix is examined.

    Returns:
        True when the lower-cased extension is a MarkItDown tabular format.
    """
    return pathlib.Path(filename).suffix.lower() in MARKITDOWN_EXTENSIONS


class MarkItDownDocumentParser:
    """Parse tabular documents (spreadsheets) via Microsoft MarkItDown.

    Satisfies ``DocumentParserPort``. Conversion is local, deterministic, and
    fast (no VLM/OCR, no network), so there is no poll loop and no per-document
    timeout — a spreadsheet that previously consumed a 30-min Docling slot now
    completes in well under a second.
    """

    def __init__(self) -> None:
        """Construct the parser, importing ``markitdown`` lazily at call time.

        The heavyweight import is deferred to ``parse`` so that importing this
        module (and the package factory) stays cheap for consumers that never
        touch a spreadsheet.
        """
        self._converter: Any | None = None

    def _get_converter(self) -> Any:
        """Return a cached ``MarkItDown`` converter, importing on first use.

        Raises:
            ImportError: If the ``markitdown`` package is not installed.
        """
        if self._converter is None:
            from markitdown import MarkItDown

            # enable_plugins=False keeps conversion deterministic and avoids
            # pulling optional third-party plugins into the parse path.
            self._converter = MarkItDown(enable_plugins=False)
        return self._converter

    def parse(
        self,
        file_content: bytes,
        filename: str,
    ) -> list[dict[str, Any]]:
        """Parse a spreadsheet and return normalized element dicts.

        MarkItDown renders each worksheet to GitHub-flavored Markdown
        (tables as pipe tables). The full Markdown is returned as a single
        ``Table`` element so the chunker treats it as structured tabular
        content; ``metadata.text_as_html`` carries the same content for
        downstream consumers that prefer HTML.

        Args:
            file_content: Raw bytes of the spreadsheet.
            filename: Original filename including extension.

        Returns:
            A list with a single ``Table`` element dict. Empty list when the
            spreadsheet has no extractable content.

        Raises:
            MarkItDownConversionError: On a terminal conversion failure — the
                runner marks the doc FAILED and continues (does not fail the KB).
        """
        ext = pathlib.Path(filename).suffix.lower()
        # Gate the memory-heavy conversion behind the module-level semaphore so
        # at most ``MARKITDOWN_MAX_CONCURRENCY`` workbooks inflate in memory at
        # once, independent of the runner's document-level concurrency.
        with _PARSE_SEMAPHORE:
            if ext in _EXCEL_EXTENSIONS:
                markdown, parser = self._parse_excel_streaming(file_content, filename, ext)
            else:
                markdown, parser = self._parse_with_markitdown(file_content, filename, ext)

        markdown = sanitize_text((markdown or '').strip())
        if not markdown:
            logger.warning('%s produced no content for %s', parser, filename)
            return []

        logger.info(
            '%s parsed %s: %d chars of markdown (no Docling/VLM)',
            parser,
            filename,
            len(markdown),
        )
        return [
            {
                'type': 'Table',
                'text': markdown,
                'metadata': {
                    'filename': filename,
                    'text_as_html': markdown,
                    'parser': parser,
                },
            }
        ]

    def _parse_with_markitdown(
        self,
        file_content: bytes,
        filename: str,
        ext: str,
    ) -> tuple[str, str]:
        """Convert *file_content* via MarkItDown (whole-stream load).

        Used for ``.csv`` and as the fallback when streaming an Excel workbook
        fails. Returns ``(markdown, 'markitdown')``.

        Raises:
            ImportError: If ``markitdown`` is not installed (operational error).
            MarkItDownConversionError: On a terminal conversion failure.
        """
        try:
            converter = self._get_converter()
            # MarkItDown infers type from the stream; passing the extension
            # hint makes detection deterministic for ambiguous byte streams.
            result = converter.convert_stream(
                BytesIO(file_content),
                file_extension=ext,
            )
        except ImportError:
            # Missing dependency is an operational/config error, not a bad
            # document — re-raise so it surfaces loudly rather than silently
            # fail-emptying every spreadsheet.
            raise
        except Exception as exc:  # noqa: BLE001 — normalize to terminal error
            raise MarkItDownConversionError(
                f'MarkItDown failed to convert {filename}: {exc}'
            ) from exc
        return (result.text_content or ''), 'markitdown'

    def _parse_excel_streaming(
        self,
        file_content: bytes,
        filename: str,
        ext: str,
    ) -> tuple[str, str]:
        """Stream an Excel workbook to Markdown with bounded memory.

        Unlike MarkItDown's ``convert_stream`` (which loads the entire
        workbook and builds one large object graph), this reads each sheet via
        ``openpyxl.load_workbook(read_only=True, data_only=True)`` and iterates
        rows with ``iter_rows``, appending pipe-table Markdown incrementally.
        openpyxl's read-only path is lazy and constant-memory, so peak memory
        is bounded to a small row window plus the growing output string, rather
        than the whole sheet materialized at once.

        ``read_only=True`` uses openpyxl's lazy, constant-memory read path;
        ``data_only=True`` returns cached cell values (not formulas). On any
        failure (including a non-``.xlsx`` file openpyxl cannot read, e.g. a
        legacy ``.xls``) it falls back to MarkItDown so behavior is never worse
        than before.

        Returns ``(markdown, 'openpyxl-stream')`` on success, or the MarkItDown
        result on fallback.

        Raises:
            MarkItDownConversionError: Only if BOTH streaming and the MarkItDown
                fallback fail terminally.
        """
        try:
            import openpyxl
        except ImportError:
            # openpyxl ships with markitdown[xlsx]; if it is somehow absent,
            # fall back to MarkItDown rather than failing the document.
            logger.warning('openpyxl unavailable — falling back to MarkItDown for %s', filename)
            return self._parse_with_markitdown(file_content, filename, ext)

        # openpyxl only reads the OOXML ``.xlsx`` format. Legacy ``.xls`` is a
        # different binary format — delegate those straight to MarkItDown.
        if ext != '.xlsx':
            return self._parse_with_markitdown(file_content, filename, ext)

        try:
            workbook = openpyxl.load_workbook(
                BytesIO(file_content),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:  # noqa: BLE001 — fall back, don't fail yet
            logger.warning(
                'openpyxl could not open %s (%s) — falling back to MarkItDown',
                filename,
                exc,
            )
            return self._parse_with_markitdown(file_content, filename, ext)

        try:
            parts: list[str] = []
            for sheet in workbook.worksheets:
                rendered = self._render_sheet_markdown(sheet)
                if rendered:
                    parts.append(f'## {sheet.title}\n\n{rendered}')
            markdown = '\n\n'.join(parts)
        except Exception as exc:  # noqa: BLE001 — normalize to terminal error
            raise MarkItDownConversionError(
                f'openpyxl streaming failed for {filename}: {exc}'
            ) from exc
        finally:
            # read_only workbooks hold an open zip handle — always close it.
            workbook.close()

        return markdown, 'openpyxl-stream'

    @staticmethod
    def _render_sheet_markdown(sheet: Any) -> str:
        """Render one worksheet to a GitHub-flavored Markdown pipe table.

        Streams rows via ``iter_rows`` so only a bounded window of cells is
        materialized at a time. Trailing fully-empty rows/columns are trimmed.
        The first non-empty row is treated as the header (matching MarkItDown's
        spreadsheet rendering). Returns ``''`` for an empty sheet.

        Args:
            sheet: An openpyxl read-only worksheet.

        Returns:
            Markdown table text, or ``''`` when the sheet has no cells.
        """

        def _fmt(value: Any) -> str:
            if value is None:
                return ''
            # Escape pipes so cell content never breaks the table structure.
            return str(value).replace('\n', ' ').replace('|', '\\|').strip()

        lines: list[str] = []
        header_written = False
        ncols = 0
        for row in sheet.iter_rows(values_only=True):
            cells = [_fmt(v) for v in row]
            # Skip fully-empty rows to avoid emitting blank table lines.
            if not any(cells):
                continue
            if not header_written:
                ncols = len(cells)
                lines.append('| ' + ' | '.join(cells) + ' |')
                lines.append('| ' + ' | '.join(['---'] * ncols) + ' |')
                header_written = True
                continue
            # Pad/truncate to the header width so the table stays rectangular.
            if len(cells) < ncols:
                cells = cells + [''] * (ncols - len(cells))
            elif len(cells) > ncols:
                cells = cells[:ncols]
            lines.append('| ' + ' | '.join(cells) + ' |')
        return '\n'.join(lines)

    def health_check(self) -> dict[str, Any]:
        """Report whether the ``markitdown`` dependency is importable.

        Returns:
            ``{'status': 'healthy', 'parser': 'markitdown'}`` when importable,
            otherwise ``{'status': 'unhealthy', 'error': ...}``. Never raises.
        """
        try:
            self._get_converter()
        except Exception as exc:  # noqa: BLE001 — health checks never raise
            return {'status': 'unhealthy', 'parser': 'markitdown', 'error': str(exc)}
        return {'status': 'healthy', 'parser': 'markitdown'}
